"""
Wallet statement export: PDF (ReportLab) and Excel (openpyxl).
"""
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

from sqlalchemy.orm import Session

from models.wallet import ClientWallet, Transaction
from models.contact import Contact


def _get_wallet_data(db: Session, org_id: str, wallet_id: str, date_from: Optional[date] = None, date_to: Optional[date] = None):
    """Fetch wallet + filtered transactions for export."""
    wallet = db.query(ClientWallet).filter(
        ClientWallet.id == wallet_id,
        ClientWallet.org_id == org_id,
    ).first()
    if not wallet:
        return None, []
    contact = db.query(Contact).filter(Contact.id == wallet.contact_id).first()
    q = db.query(Transaction).filter(Transaction.wallet_id == wallet_id).order_by(Transaction.created_at.asc())
    if date_from:
        q = q.filter(Transaction.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.filter(Transaction.created_at <= datetime.combine(date_to, datetime.max.time()))
    txns = q.all()
    return {
        "wallet": wallet,
        "contact": contact,
        "contact_name": contact.name if contact else "Unknown",
        "currency": wallet.currency or "AED",
        "balance": wallet.balance,
    }, txns


def generate_statement_pdf(db: Session, org_id: str, wallet_id: str, date_from: Optional[date] = None, date_to: Optional[date] = None) -> Optional[bytes]:
    """Generate a PDF wallet statement. Returns bytes or None if wallet not found."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    info, txns = _get_wallet_data(db, org_id, wallet_id, date_from, date_to)
    if info is None:
        return None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Client Wallet Statement", styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"<b>Client:</b> {info['contact_name']}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Currency:</b> {info['currency']}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Current Balance:</b> {info['balance']:,.2f} {info['currency']}", styles["Normal"]))
    if date_from or date_to:
        period = f"{date_from or 'Start'} to {date_to or 'Now'}"
        elements.append(Paragraph(f"<b>Period:</b> {period}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    if txns:
        header = ["Date", "Type", "Description", "Amount", "Balance After"]
        rows = [header]
        for t in txns:
            rows.append([
                t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
                t.type.value if t.type else "",
                (t.description or "")[:60],
                f"{t.amount:,.2f}",
                f"{t.balance_after:,.2f}",
            ])
        table = Table(rows, colWidths=[90, 80, 180, 80, 80])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a365d")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
            ("ALIGN", (3, 0), (4, -1), "RIGHT"),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("No transactions in this period.", styles["Normal"]))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))

    doc.build(elements)
    return buf.getvalue()


def generate_statement_excel(db: Session, org_id: str, wallet_id: str, date_from: Optional[date] = None, date_to: Optional[date] = None) -> Optional[bytes]:
    """Generate an Excel wallet statement. Returns bytes or None if wallet not found."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    info, txns = _get_wallet_data(db, org_id, wallet_id, date_from, date_to)
    if info is None:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Wallet Statement"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")

    ws.append(["Client Wallet Statement"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Client:", info["contact_name"]])
    ws.append(["Currency:", info["currency"]])
    ws.append(["Current Balance:", float(info["balance"])])
    if date_from or date_to:
        ws.append(["Period:", f"{date_from or 'Start'} to {date_to or 'Now'}"])
    ws.append([])

    headers = ["Date", "Type", "Description", "Amount", "VAT", "Total", "Balance After", "Reference"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for t in txns:
        ws.append([
            t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
            t.type.value if t.type else "",
            t.description or "",
            float(t.amount_exclusive) if t.amount_exclusive else float(t.amount),
            float(t.vat_amount) if t.vat_amount else 0,
            float(t.amount),
            float(t.balance_after),
            t.reference_id or "",
        ])

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["C"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
