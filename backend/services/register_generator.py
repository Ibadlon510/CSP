"""
Generate UAE-style registers: UBO, Partners, Directors (PDF and Excel).
Stores files under uploads/registers/{org_id}/ and creates ComplianceSnapshot.
"""
import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Border, Side

from models.contact import Contact, ContactType, ContactAddress
from models.compliance import OwnershipLink, OwnershipLinkType, ComplianceSnapshot, RegisterType
from services.ubo_resolver import resolve_ubos


def _version_hash(entity_contact_id: str, snapshot_data: list) -> str:
    raw = entity_contact_id + "|" + str(sorted(str(x) for x in snapshot_data))
    return hashlib.sha256(raw.encode()).hexdigest()[:16]


def _get_entity_address(db: Session, contact_id: str) -> str:
    addr = (
        db.query(ContactAddress)
        .filter(ContactAddress.contact_id == contact_id, ContactAddress.is_primary == True)
        .first()
    )
    if not addr:
        addr = db.query(ContactAddress).filter(ContactAddress.contact_id == contact_id).first()
    if not addr:
        return ""
    parts = [addr.address_line_1, addr.address_line_2, addr.city, addr.state_emirate, addr.country]
    return ", ".join(p for p in parts if p)


def _build_ubo_data(db: Session, org_id: str, entity_contact_id: str) -> tuple[list[dict], Optional[str]]:
    entity = db.query(Contact).filter(Contact.id == entity_contact_id, Contact.org_id == org_id).first()
    if not entity:
        return [], None
    senior_id = getattr(entity, "senior_manager_contact_id", None)
    result = resolve_ubos(db, org_id, entity_contact_id, senior_manager_contact_id=senior_id)
    rows = []
    for u in result["ubos"]:
        c = db.query(Contact).filter(Contact.id == u["contact_id"], Contact.org_id == org_id).first()
        if not c:
            continue
        rows.append({
            "name": c.name,
            "nationality": c.nationality or "",
            "passport_no": c.passport_no or "",
            "date_of_birth": str(c.date_of_birth) if c.date_of_birth else "",
            "address": _get_entity_address(db, c.id),
            "effective_pct": u["effective_pct"],
            "is_control": u.get("is_control", False),
            "is_senior_manager_fallback": u.get("is_senior_manager_fallback", False),
        })
    return rows, entity.name


def _build_partners_data(db: Session, org_id: str, entity_contact_id: str) -> tuple[list[dict], Optional[str]]:
    entity = db.query(Contact).filter(Contact.id == entity_contact_id, Contact.org_id == org_id).first()
    if not entity:
        return [], None
    links = (
        db.query(OwnershipLink)
        .filter(
            OwnershipLink.org_id == org_id,
            OwnershipLink.owned_contact_id == entity_contact_id,
            OwnershipLink.link_type == OwnershipLinkType.OWNERSHIP,
        )
        .all()
    )
    rows = []
    for l in links:
        c = db.query(Contact).filter(Contact.id == l.owner_contact_id, Contact.org_id == org_id).first()
        if not c:
            continue
        rows.append({
            "name": c.name,
            "contact_type": c.contact_type.value if c.contact_type else "company",
            "nationality": c.nationality or c.country or "",
            "passport_no": c.passport_no or "",
            "address": _get_entity_address(db, c.id),
            "percentage": l.percentage or 0,
        })
    return rows, entity.name


def _build_directors_data(db: Session, org_id: str, entity_contact_id: str) -> tuple[list[dict], Optional[str]]:
    entity = db.query(Contact).filter(Contact.id == entity_contact_id, Contact.org_id == org_id).first()
    if not entity:
        return [], None
    links = (
        db.query(OwnershipLink)
        .filter(
            OwnershipLink.org_id == org_id,
            OwnershipLink.owned_contact_id == entity_contact_id,
            OwnershipLink.link_type.in_([OwnershipLinkType.DIRECTOR, OwnershipLinkType.MANAGES]),
        )
        .all()
    )
    rows = []
    for l in links:
        c = db.query(Contact).filter(Contact.id == l.owner_contact_id, Contact.org_id == org_id).first()
        if not c:
            continue
        rows.append({
            "name": c.name,
            "nationality": c.nationality or "",
            "passport_no": c.passport_no or "",
            "designation": getattr(c, "designation_title", None) or "Director",
            "is_nominee": l.is_nominee == "true",
        })
    if not rows:
        return [], entity.name
    return rows, entity.name


def _generate_ubo_pdf(file_path: str, entity_name: str, rows: list[dict], register_title: str) -> None:
    doc = SimpleDocTemplate(file_path, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = [Paragraph(f"<b>Register of Beneficial Owners</b>", styles["Heading1"]), Spacer(1, 6)]
    story.append(Paragraph("UAE Cabinet Decision No. (109) of 2023", styles["Normal"]))
    story.append(Paragraph(f"<b>Entity:</b> {entity_name}", styles["Normal"]))
    story.append(Paragraph(f"<b>Generated:</b> {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    story.append(Spacer(1, 20))
    headers = ["No.", "Full Name", "Nationality", "Passport No.", "DOB", "Address", "Effective %", "Control / SM"]
    data = [headers]
    for i, r in enumerate(rows, 1):
        ctrl = "Control" if r.get("is_control") else ("Senior Manager" if r.get("is_senior_manager_fallback") else "")
        data.append([
            str(i), r.get("name", ""), r.get("nationality", ""), r.get("passport_no", ""),
            r.get("date_of_birth", ""), (r.get("address", "") or "")[:50], str(r.get("effective_pct", "")), ctrl,
        ])
    if not rows:
        data.append(["No UBOs identified", "", "", "", "", "", "", ""])
    t = Table(data, colWidths=[1*cm, 3*cm, 2*cm, 2.5*cm, 2*cm, 4*cm, 1.5*cm, 2*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(t)
    doc.build(story)


def _generate_ubo_excel(file_path: str, entity_name: str, rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Register of UBOs"
    ws.append(["Register of Beneficial Owners"])
    ws.append(["Entity:", entity_name])
    ws.append(["Generated:", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    ws.append([])
    headers = ["No.", "Full Name", "Nationality", "Passport No.", "DOB", "Address", "Effective %", "Control / SM"]
    ws.append(headers)
    for i, r in enumerate(rows, 1):
        ctrl = "Control" if r.get("is_control") else ("Senior Manager" if r.get("is_senior_manager_fallback") else "")
        ws.append([
            i, r.get("name", ""), r.get("nationality", ""), r.get("passport_no", ""),
            r.get("date_of_birth", ""), r.get("address", ""), r.get("effective_pct", ""), ctrl,
        ])
    thin = Side(style="thin")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb.save(file_path)


def _generate_partners_pdf(file_path: str, entity_name: str, rows: list[dict]) -> None:
    doc = SimpleDocTemplate(file_path, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = [Paragraph("<b>Register of Partners / Members</b>", styles["Heading1"]), Spacer(1, 12)]
    story.append(Paragraph(f"<b>Entity:</b> {entity_name}", styles["Normal"]))
    story.append(Spacer(1, 20))
    headers = ["No.", "Name", "Type", "Nationality", "Passport/License", "Address", "Percentage"]
    data = [headers]
    for i, r in enumerate(rows, 1):
        data.append([str(i), r.get("name", ""), r.get("contact_type", ""), r.get("nationality", ""), r.get("passport_no", ""), (r.get("address", "") or "")[:40], str(r.get("percentage", ""))])
    if not rows:
        data.append(["No partners", "", "", "", "", "", ""])
    t = Table(data, colWidths=[1*cm, 3.5*cm, 1.5*cm, 2*cm, 2.5*cm, 4*cm, 1.5*cm])
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("FONTSIZE", (0, 0), (-1, -1), 8), ("GRID", (0, 0), (-1, -1), 0.5, colors.grey)]))
    story.append(t)
    doc.build(story)


def _generate_partners_excel(file_path: str, entity_name: str, rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Partners"
    ws.append(["Register of Partners / Members"])
    ws.append(["Entity:", entity_name])
    ws.append([])
    headers = ["No.", "Name", "Type", "Nationality", "Passport/License", "Address", "Percentage"]
    ws.append(headers)
    for i, r in enumerate(rows, 1):
        ws.append([i, r.get("name", ""), r.get("contact_type", ""), r.get("nationality", ""), r.get("passport_no", ""), r.get("address", ""), r.get("percentage", "")])
    wb.save(file_path)


def _generate_directors_pdf(file_path: str, entity_name: str, rows: list[dict]) -> None:
    doc = SimpleDocTemplate(file_path, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = [Paragraph("<b>Register of Directors / Managers</b>", styles["Heading1"]), Spacer(1, 12)]
    story.append(Paragraph(f"<b>Entity:</b> {entity_name}", styles["Normal"]))
    story.append(Spacer(1, 20))
    headers = ["No.", "Name", "Nationality", "Passport No.", "Designation", "Nominee"]
    data = [headers]
    for i, r in enumerate(rows, 1):
        data.append([str(i), r.get("name", ""), r.get("nationality", ""), r.get("passport_no", ""), r.get("designation", ""), "Yes" if r.get("is_nominee") else "No"])
    if not rows:
        data.append(["No directors", "", "", "", "", ""])
    t = Table(data, colWidths=[1*cm, 4*cm, 2.5*cm, 3*cm, 3*cm, 2*cm])
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("FONTSIZE", (0, 0), (-1, -1), 8), ("GRID", (0, 0), (-1, -1), 0.5, colors.grey)]))
    story.append(t)
    doc.build(story)


def _generate_directors_excel(file_path: str, entity_name: str, rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Directors"
    ws.append(["Register of Directors / Managers"])
    ws.append(["Entity:", entity_name])
    ws.append([])
    headers = ["No.", "Name", "Nationality", "Passport No.", "Designation", "Nominee"]
    ws.append(headers)
    for i, r in enumerate(rows, 1):
        ws.append([i, r.get("name", ""), r.get("nationality", ""), r.get("passport_no", ""), r.get("designation", ""), "Yes" if r.get("is_nominee") else "No"])
    wb.save(file_path)


def generate_register(
    db: Session,
    org_id: str,
    entity_contact_id: str,
    register_type: str,
    fmt: str,
    generated_by: Optional[str] = None,
) -> ComplianceSnapshot:
    """Generate register (UBO, partners, or directors) as PDF or Excel; save file and create snapshot."""
    reg = register_type.lower()
    if reg not in ("ubo", "partners", "directors"):
        raise ValueError("register_type must be ubo, partners, or directors")
    if fmt.lower() not in ("pdf", "excel"):
        raise ValueError("format must be pdf or excel")
    ext = "pdf" if fmt.lower() == "pdf" else "xlsx"
    base = Path(__file__).resolve().parent.parent
    uploads_dir = base / "uploads" / "registers" / org_id
    uploads_dir.mkdir(parents=True, exist_ok=True)
    now = datetime.now(timezone.utc)
    snapshot_id = None  # we'll create after we have file_path

    if reg == "ubo":
        rows, entity_name = _build_ubo_data(db, org_id, entity_contact_id)
        entity_name = entity_name or "Unknown"
        version = _version_hash(entity_contact_id, [r.get("name") for r in rows])
        filename = f"ubo_{entity_contact_id[:8]}_{now.strftime('%Y%m%d%H%M')}.{ext}"
        file_path = str(uploads_dir / filename)
        rel_path = f"uploads/registers/{org_id}/{filename}"
        if fmt.lower() == "pdf":
            _generate_ubo_pdf(file_path, entity_name, rows, "Register of Beneficial Owners")
        else:
            _generate_ubo_excel(file_path, entity_name, rows)
    elif reg == "partners":
        rows, entity_name = _build_partners_data(db, org_id, entity_contact_id)
        entity_name = entity_name or "Unknown"
        version = _version_hash(entity_contact_id, [r.get("name") for r in rows])
        filename = f"partners_{entity_contact_id[:8]}_{now.strftime('%Y%m%d%H%M')}.{ext}"
        file_path = str(uploads_dir / filename)
        rel_path = f"uploads/registers/{org_id}/{filename}"
        if fmt.lower() == "pdf":
            _generate_partners_pdf(file_path, entity_name, rows)
        else:
            _generate_partners_excel(file_path, entity_name, rows)
    else:
        rows, entity_name = _build_directors_data(db, org_id, entity_contact_id)
        entity_name = entity_name or "Unknown"
        version = _version_hash(entity_contact_id, [r.get("name") for r in rows])
        filename = f"directors_{entity_contact_id[:8]}_{now.strftime('%Y%m%d%H%M')}.{ext}"
        file_path = str(uploads_dir / filename)
        rel_path = f"uploads/registers/{org_id}/{filename}"
        if fmt.lower() == "pdf":
            _generate_directors_pdf(file_path, entity_name, rows)
        else:
            _generate_directors_excel(file_path, entity_name, rows)

    rt = RegisterType.UBO if reg == "ubo" else (RegisterType.PARTNERS if reg == "partners" else RegisterType.DIRECTORS)
    snapshot = ComplianceSnapshot(
        org_id=org_id,
        entity_contact_id=entity_contact_id,
        register_type=rt,
        version_hash=version,
        file_path=rel_path,
        generated_at=now,
        generated_by=generated_by,
        snapshot_data=rows,
    )
    db.add(snapshot)
    db.commit()
    db.refresh(snapshot)
    return snapshot
